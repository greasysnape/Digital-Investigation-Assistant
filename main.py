"""
main_react.py
=============
DIA — Digital Investigation Assistant (ReACT version)

Uses LangGraph create_react_agent with ChatOllama for local inference.
All agent personas merged into one system prompt.

Run:
    streamlit run main_react.py
"""

import base64
import glob
import json
import os
import re
import shutil
import time
import uuid
from datetime import datetime
from typing import Optional

import streamlit as st

from langchain_core.messages import HumanMessage

from config import (
    ensure_directories, HISTORY_DIR,
    get_available_cases,
)
from tools import (
    load_case_db, unload_case_db, set_active_session,
    get_embedder, get_shared_dbs,
)
from react_engine import create_investigation_agent, DEFAULT_MODEL

try:
    from eval.eval_logger import log_team as _log_team
except Exception:
    _log_team = None

# ─────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────
MAX_STEPS   = 12
ATTACHMENTS_DIR = "attachments"


def _get_ollama_models() -> list:
    """Fetch available model names from the local Ollama server."""
    import requests as _req
    try:
        r = _req.get("http://localhost:11434/api/tags", timeout=5)
        r.raise_for_status()
        return sorted([m["name"] for m in r.json().get("models", [])])
    except Exception:
        return [DEFAULT_MODEL]

st.set_page_config(
    page_title="Digital Forensic Assistant",
    page_icon="🕵️",
    layout="wide"
)
ensure_directories()
os.makedirs(ATTACHMENTS_DIR, exist_ok=True)

# Pre-warm FAISS caches in the main Streamlit thread.
get_embedder()
get_shared_dbs()


# ─────────────────────────────────────────────────────────────
# File handling helpers
# ─────────────────────────────────────────────────────────────
SUPPORTED_TEXT_EXTS = {".txt", ".csv", ".json", ".md", ".log", ".xml", ".html"}
SUPPORTED_DOC_EXTS = {".pdf", ".docx", ".xlsx", ".xls"}
SUPPORTED_IMG_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
ALL_SUPPORTED_EXTS = SUPPORTED_TEXT_EXTS | SUPPORTED_DOC_EXTS | SUPPORTED_IMG_EXTS


def _save_attachment(uploaded_file, session_id: str) -> dict:
    """Save an uploaded file to attachments/{session_id}/ and return metadata."""
    session_dir = os.path.join(ATTACHMENTS_DIR, session_id[:8])
    os.makedirs(session_dir, exist_ok=True)

    # Unique filename to avoid collisions
    file_id = uuid.uuid4().hex[:8]
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", uploaded_file.name)
    save_name = f"{file_id}_{safe_name}"
    save_path = os.path.join(session_dir, save_name)

    with open(save_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    return {
        "file_id": file_id,
        "original_name": uploaded_file.name,
        "saved_path": save_path,
        "size_bytes": uploaded_file.size,
        "mime_type": uploaded_file.type or "",
        "ext": os.path.splitext(uploaded_file.name)[1].lower(),
    }


def _extract_text(file_meta: dict) -> str:
    """Extract text content from a file for inclusion in the prompt."""
    ext = file_meta["ext"]
    path = file_meta["saved_path"]

    try:
        # Plain text files
        if ext in SUPPORTED_TEXT_EXTS:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()[:10000]

        # PDF
        if ext == ".pdf":
            try:
                import pdfplumber
                text_parts = []
                with pdfplumber.open(path) as pdf:
                    for page in pdf.pages[:30]:
                        t = page.extract_text()
                        if t:
                            text_parts.append(t)
                return "\n".join(text_parts)[:10000]
            except ImportError:
                return "[PDF support requires pdfplumber: pip install pdfplumber]"

        # DOCX
        if ext == ".docx":
            try:
                from docx import Document
                doc = Document(path)
                return "\n".join(p.text for p in doc.paragraphs)[:10000]
            except ImportError:
                return "[DOCX support requires python-docx: pip install python-docx]"

        # Excel
        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                rows = []
                for sheet in wb.sheetnames[:3]:
                    ws = wb[sheet]
                    rows.append(f"[Sheet: {sheet}]")
                    for row in ws.iter_rows(max_row=200, values_only=True):
                        rows.append(",".join(str(c) if c is not None else "" for c in row))
                wb.close()
                return "\n".join(rows)[:10000]
            except ImportError:
                return "[Excel support requires openpyxl: pip install openpyxl]"

        return f"[Unsupported file type: {ext}]"

    except Exception as e:
        return f"[Error reading file: {e}]"


def _encode_image_base64(file_meta: dict) -> str:
    """Encode an image file to base64 for multimodal LLM input."""
    with open(file_meta["saved_path"], "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def _build_message_content(prompt: str, attachments: list) -> list:
    """
    Build a HumanMessage content list with text and optional images.
    For text attachments, content is prepended to the prompt.
    For images, they are added as image_url parts.
    """
    text_parts = []
    image_parts = []

    for att in attachments:
        ext = att["ext"]
        if ext in SUPPORTED_IMG_EXTS:
            b64 = _encode_image_base64(att)
            mime = att["mime_type"] or f"image/{ext.lstrip('.')}"
            image_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}"},
            })
        else:
            extracted = _extract_text(att)
            text_parts.append(f"[Attached file: {att['original_name']}]\n{extracted}")

    # Build the content list
    combined_text = prompt
    if text_parts:
        combined_text = "\n\n".join(text_parts) + f"\n\n---\nUser query: {prompt}"

    content = [{"type": "text", "text": combined_text}]
    content.extend(image_parts)

    return content


# ─────────────────────────────────────────────────────────────
# History helpers
# ─────────────────────────────────────────────────────────────
def _history_dir(case: Optional[str]) -> str:
    sub = case if case else "No_Case"
    path = os.path.join(HISTORY_DIR, sub)
    os.makedirs(path, exist_ok=True)
    return path

def _session_file(session_id: str, title: str, case: Optional[str]) -> str:
    safe = re.sub(r'[\\/*?:"<>|]', "", title)
    return os.path.join(_history_dir(case), f"{safe}_{session_id[:8]}.json")

def load_all_sessions() -> dict:
    sessions = {}
    for fp in glob.glob(os.path.join(HISTORY_DIR, "**", "*.json"), recursive=True):
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
            key = data.get("id", os.path.basename(fp).replace(".json", ""))
            sessions[key] = data
        except Exception:
            continue
    return sessions

def save_session(session: dict):
    fp = _session_file(session["id"], session["title"], session.get("case"))
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(session, f, ensure_ascii=False, indent=4)

def delete_session_file(session: dict):
    fp = _session_file(session["id"], session["title"], session.get("case"))
    if os.path.exists(fp):
        os.remove(fp)


# ─────────────────────────────────────────────────────────────
# Session management
# ─────────────────────────────────────────────────────────────
if "sessions" not in st.session_state:
    st.session_state.sessions = load_all_sessions()

def create_new_session(case: Optional[str] = None) -> str:
    new_id = str(uuid.uuid4())
    title  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session = {"id": new_id, "title": title, "case": case, "messages": []}
    st.session_state.sessions[new_id] = session
    save_session(session)
    return new_id

def get_current_session() -> dict:
    sid = st.session_state.get("current_session_id")
    if sid and sid in st.session_state.sessions:
        return st.session_state.sessions[sid]
    if st.session_state.sessions:
        most_recent = max(st.session_state.sessions.values(), key=lambda s: s["title"])
        st.session_state.current_session_id = most_recent["id"]
        return most_recent
    new_id = create_new_session(st.session_state.get("selected_case"))
    st.session_state.current_session_id = new_id
    return st.session_state.sessions[new_id]

def delete_session(sid: str):
    if sid in st.session_state.sessions:
        delete_session_file(st.session_state.sessions[sid])
        # Clean up attachments
        att_dir = os.path.join(ATTACHMENTS_DIR, sid[:8])
        if os.path.isdir(att_dir):
            shutil.rmtree(att_dir, ignore_errors=True)
        del st.session_state.sessions[sid]
        remaining = list(st.session_state.sessions.keys())
        st.session_state.current_session_id = (
            remaining[-1] if remaining
            else create_new_session(st.session_state.get("selected_case"))
        )


# ─────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🕵️ Digital Investigation Assistant")

    # ── Model Selection ──
    st.markdown("### 🤖 Model")
    ollama_models = _get_ollama_models()
    prev_model = st.session_state.get("selected_model", DEFAULT_MODEL)
    model_idx = ollama_models.index(prev_model) if prev_model in ollama_models else 0
    selected_model = st.selectbox(
        "Select LLM",
        options=ollama_models,
        index=model_idx,
        key="model_selector",
        label_visibility="collapsed",
    )
    st.session_state.selected_model = selected_model
    REACT_MODEL = selected_model

    st.markdown("---")
    st.markdown("### 📁 Case Selection")

    available_cases = get_available_cases()
    case_options    = ["— No Case —"] + available_cases
    prev_case       = st.session_state.get("selected_case")
    default_idx     = 0
    if prev_case and prev_case in available_cases:
        default_idx = available_cases.index(prev_case) + 1

    selected_option = st.selectbox(
        "Select Investigation Case",
        options=case_options,
        index=default_idx,
        key="case_selector"
    )
    selectbox_case = None if selected_option == "— No Case —" else selected_option

    # Only update from selectbox if user actually changed it (not from session switch)
    if selectbox_case != prev_case and not st.session_state.get("_case_switched_by_history"):
        st.session_state.selected_case = selectbox_case
        if selectbox_case:
            loaded = load_case_db(selectbox_case)
            st.toast("✅ Case DB loaded." if loaded else "⚠️ No case DB found.")
        else:
            unload_case_db()
        existing = [
            s for s in st.session_state.sessions.values()
            if s.get("case") == selectbox_case
        ]
        if existing:
            most_recent = max(existing, key=lambda s: s["title"])
            st.session_state.current_session_id = most_recent["id"]
        else:
            new_id = create_new_session(selectbox_case)
            st.session_state.current_session_id = new_id
        st.rerun()
    st.session_state.pop("_case_switched_by_history", None)
    selected_case = st.session_state.get("selected_case")

    if selected_case:
        case_db_dir = f"./cases/{selected_case}/case_db"
        if glob.glob(f"{case_db_dir}/*.faiss"):
            st.success(f"📦 Case DB: {selected_case}")
        else:
            st.warning(f"📂 Case: {selected_case} (no case DB)")
    else:
        st.info("📂 No case selected — using shared DBs")

    st.markdown("---")
    if st.button("➕ New Chat", type="primary", use_container_width=True):
        new_id = create_new_session(selected_case)
        st.session_state.current_session_id = new_id
        st.rerun()

    st.markdown("---")
    st.markdown("### 📂 Chat History")

    sessions_by_case: dict = {}
    for sid, sess in st.session_state.sessions.items():
        c = sess.get("case") or "No_Case"
        sessions_by_case.setdefault(c, []).append(sess)

    seen, display_order = set(), []
    def _add(key):
        if key not in seen and key in sessions_by_case:
            seen.add(key); display_order.append((key, sessions_by_case[key]))
    if selected_case:
        _add(selected_case)
    for c in sessions_by_case:
        if c != "No_Case": _add(c)
    _add("No_Case")

    for case_label, sess_list in display_order:
        st.markdown(f"**{case_label}**")
        for sess in sorted(sess_list, key=lambda s: s["title"], reverse=True):
            active = sess["id"] == st.session_state.get("current_session_id")
            label  = f"{'🟢 ' if active else ''}{sess['title']}"
            col1, col2 = st.columns([0.85, 0.15])
            if col1.button(label, key=f"s_{sess['id']}", use_container_width=True):
                st.session_state.current_session_id = sess["id"]
                sess_case = sess.get("case")
                if sess_case != st.session_state.get("selected_case"):
                    st.session_state.selected_case = sess_case
                    st.session_state._case_switched_by_history = True
                    load_case_db(sess_case) if sess_case else unload_case_db()
                st.rerun()
            if col2.button("✕", key=f"d_{sess['id']}"):
                delete_session(sess["id"])
                st.rerun()


# ─────────────────────────────────────────────────────────────
# Main Chat
# ─────────────────────────────────────────────────────────────
session    = get_current_session()
case_label = session.get("case") or "No Case"
st.header(f"🕵️ Digital Forensic Assistant — {case_label}")
st.caption(f"Session: {session['title']}  |  Model: {REACT_MODEL}")

# Display chat history
for msg in session["messages"]:
    with st.chat_message(msg["role"], avatar="🕵️" if msg["role"] == "assistant" else None):
        # Show attachment info if present
        if msg.get("attachments"):
            for att in msg["attachments"]:
                ext = att.get("ext", "")
                if ext in SUPPORTED_IMG_EXTS and os.path.exists(att.get("saved_path", "")):
                    st.image(att["saved_path"], caption=att["original_name"], width=300)
                else:
                    st.caption(f"📎 {att['original_name']} ({att.get('size_bytes', 0) // 1024}KB)")
        st.markdown(msg["content"])

if submission := st.chat_input(
    "How may I assist you with your investigation?",
    accept_file="multiple",
    file_type=[ext.lstrip(".") for ext in ALL_SUPPORTED_EXTS],
):
    prompt = submission.text if hasattr(submission, "text") else str(submission)
    uploaded_files = submission.files if hasattr(submission, "files") else []

    # Process attachments
    attachments_meta = []
    for uf in uploaded_files:
        meta = _save_attachment(uf, session["id"])
        attachments_meta.append(meta)

    # Display user message
    with st.chat_message("user"):
        for att in attachments_meta:
            if att["ext"] in SUPPORTED_IMG_EXTS:
                st.image(att["saved_path"], caption=att["original_name"], width=300)
            else:
                st.caption(f"📎 {att['original_name']} ({att['size_bytes'] // 1024}KB)")
        st.markdown(prompt)

    # Save user message with attachment metadata
    user_msg = {"role": "user", "content": prompt}
    if attachments_meta:
        user_msg["attachments"] = [
            {
                "file_id": a["file_id"],
                "original_name": a["original_name"],
                "saved_path": a["saved_path"],
                "size_bytes": a["size_bytes"],
                "ext": a["ext"],
            }
            for a in attachments_meta
        ]
    session["messages"].append(user_msg)

    # Build message content for LLM
    if attachments_meta:
        message_content = _build_message_content(prompt, attachments_meta)
    else:
        message_content = prompt

    has_case = session.get("case") is not None
    agent = create_investigation_agent(has_case=has_case, use_memory=True)
    set_active_session(session["id"])

    thread_id = session["id"]
    stream_config = {
        "recursion_limit": MAX_STEPS * 3,
        "configurable": {"thread_id": thread_id},
    }

    start_time = time.time()
    final_answer = ""
    observations = []

    with st.chat_message("assistant", avatar="🕵️"):
        status_placeholder = st.status("🕵️ Analysing...", expanded=True)

        # ── Phase 1: Information Gathering ──
        inputs = {"messages": [HumanMessage(content=message_content)], "has_case": has_case}

        for step in agent.stream(inputs, config=stream_config):
            for node_name, output in step.items():
                last_msg = output["messages"][-1]

                if node_name == "reason":
                    thought = last_msg.content or ""
                    if thought:
                        status_placeholder.write(f"💭 **Thought:** {thought[:200]}...")
                    if "ANSWER:" in thought:
                        final_answer = thought

                elif node_name == "act":
                    if hasattr(last_msg, "tool_calls") and last_msg.tool_calls:
                        for tool_call in last_msg.tool_calls:
                            status_placeholder.write(f"🔧 **Action:** `{tool_call['name']}`")
                    elif last_msg.content and "DIRECT_ANSWER:" in str(last_msg.content):
                        final_answer = str(last_msg.content).split("DIRECT_ANSWER:", 1)[1].strip()

                elif node_name == "observation":
                    for tool_msg in output["messages"]:
                        full_content = str(tool_msg.content)
                        observations.append(full_content)
                        snippet = full_content[:300].replace("\n", " ")
                        status_placeholder.write(f"✅ **Observation:** {snippet}...")

        # ── Phase 2: Synthesis ──
        if observations:
            status_placeholder.write("📝 **Synthesising final answer...**")
            from react_engine import _synthesise
            final_answer = _synthesise(prompt, observations, REACT_MODEL, verbose=False)
        else:
            # No tool calls — use raw model output, clean up if needed
            if "ANSWER:" in final_answer:
                final_answer = final_answer.split("ANSWER:", 1)[1].strip()
            import re as _re
            for token in ["<|endoftext|>", "<|im_start|>", "<|im_end|>"]:
                if token in final_answer:
                    final_answer = final_answer[:final_answer.index(token)]
            final_answer = _re.sub(r"<think>.*?</think>", "", final_answer, flags=_re.DOTALL)
            if "</think>" in final_answer:
                final_answer = final_answer.split("</think>")[-1]
            final_answer = final_answer.strip()

        status_placeholder.update(label="✅ Analysis complete", state="complete", expanded=False)
        st.markdown(final_answer)

    elapsed = round(time.time() - start_time, 1)
    st.caption(f"⏱️ Time elapsed: {elapsed}s")
    session["messages"].append({"role": "assistant", "content": final_answer})
    save_session(session)

    if _log_team:
        try:
            _log_team(
                query=prompt,
                elapsed_seconds=elapsed,
                report=final_answer,
                session_id=session["id"],
            )
        except Exception:
            pass
